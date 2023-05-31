# import openpyxl
# import pandas as pd
# import streamlit as st
#
# st.set_page_config(layout="wide")
#
#
# # Function to format empty cells
# def format_empty_cells(val):
#     if pd.isnull(val):
#         return '----------'
#     return val
#
#
# with open('design.css') as source_design:
#     st.markdown(f"<style>{source_design.read()}</style>", unsafe_allow_html=True)
#
# # Load the Excel sheet
# FILE_PATH = "اجمالى الحديد والأسمنت.xlsx"
# workbook = openpyxl.load_workbook(FILE_PATH)
# sheet_names = workbook.sheetnames
#
# # Make the Sidebar
# st.sidebar.header('اجمالى توريدات الحديد والأسمنت')
# type_of_imports = st.sidebar.selectbox('اختر نوع التوريد', options=sheet_names)
#
# #
# if type_of_imports in sheet_names:
#     df = pd.read_excel(FILE_PATH, sheet_name=type_of_imports, skiprows=4)
#     select_columns = st.multiselect('اختر العمود المراد عمل فلتر عليه', options=df.columns)
#     filtered_df = df.copy()
#     for col in select_columns:
#         selected_values = st.multiselect(f' اختار{col}', options=filtered_df[col].unique())
#         filtered_df = filtered_df[filtered_df[col].isin(selected_values)]
#     formatted_df = filtered_df.applymap(format_empty_cells)
#     st.dataframe(formatted_df, width=1200)
#     total_of_quantity = filtered_df['وزن البوليصة '].sum()
#     st.dataframe({'اجمالى الكمية الموردة': [total_of_quantity]}, width=600)
#

import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(layout="wide")


# Function to format empty cells
def format_empty_cells(val):
    if pd.isnull(val):
        return '----------'
    return val


with open('design.css') as source_design:
    st.markdown(f"<style>{source_design.read()}</style>", unsafe_allow_html=True)

# Load the Excel sheet
FILE_PATH = "اجمالى الحديد والأسمنت.xlsx"
workbook = openpyxl.load_workbook(FILE_PATH)
sheet_names = workbook.sheetnames

# Make the Sidebar
st.sidebar.header('اجمالى توريدات الحديد والأسمنت')
type_of_imports = st.sidebar.selectbox('اختر نوع التوريد', options=sheet_names)

if type_of_imports in sheet_names:
    df = pd.read_excel(FILE_PATH, sheet_name=type_of_imports, skiprows=4)
    select_columns = st.multiselect('اختر العمود المراد عمل فلتر عليه', options=df.columns)
    filtered_df = df.copy()
    for col in select_columns:
        selected_values = st.multiselect(f' اختار{col}', options=filtered_df[col].unique())
        filtered_df = filtered_df[filtered_df[col].isin(selected_values)]
    formatted_df = filtered_df.applymap(format_empty_cells)

    # Apply center alignment to the dataframe
    formatted_df_styled = formatted_df.style.set_properties(**{'text-align': 'center'})

    # Display the formatted dataframe
    st.dataframe(formatted_df_styled, width=1200)

    total_of_quantity = filtered_df['وزن البوليصة '].sum()
    st.dataframe({'اجمالى الكمية الموردة': [total_of_quantity]}, width=600)

