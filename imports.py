# import pandas as pd
# import openpyxl
# import streamlit as st
#
# st.set_page_config(layout='wide')
# # load Excel sheet
# FILEPATH = "اجمالى الحديد والأسمنت.xlsx"
# work_book = openpyxl.load_workbook(FILEPATH)
# sheet_names = work_book.sheetnames
# print(sheet_names)
#
# st.sidebar.header('واردات الحديد والأسمنت')
#
# type_of_imports = st.sidebar.selectbox('اختر نوع التوريد', options=sheet_names)
#
# if type_of_imports in sheet_names:
#     df = pd.read_excel(FILEPATH, sheet_name=type_of_imports, skiprows=4)
#     column_names = df.columns.tolist()
#     filter_columns = st.multiselect('اختار العمود الذى تريد عمل فلتر عليه', options=column_names)
#     filtered_df = df.copy()
#     for col in filter_columns:
#         selected_values = st.multiselect(f' اختار{col}', options=filtered_df[col].unique())
#         filtered_df = filtered_df[filtered_df[col].isin(selected_values)]
#     st.write(filtered_df)
#     total_of_material = filtered_df['وزن البوليصة '].sum()
#     st.dataframe({'اجمالى الكمية الموردة': [total_of_material]})

# import pandas as pd
# import openpyxl
# import streamlit as st
#
# st.set_page_config(layout='wide')
#
# # Load Excel sheet
# FILEPATH = "اجمالى الحديد والأسمنت.xlsx"
# work_book = openpyxl.load_workbook(FILEPATH)
# sheet_names = work_book.sheetnames
#
# st.sidebar.header('واردات الحديد والأسمنت')
#
# type_of_imports = st.sidebar.selectbox('اختر نوع التوريد', options=sheet_names)
#
# if type_of_imports in sheet_names:
#     df = pd.read_excel(FILEPATH, sheet_name=type_of_imports, skiprows=4)
#     column_names = df.columns.tolist()
#     filter_columns = st.multiselect('اختار العمود الذى تريد عمل فلتر عليه', options=column_names)
#
#     filtered_df = df.copy()
#     for col in filter_columns:
#         selected_values = st.multiselect(f'اختر {col}', options=filtered_df[col].unique())
#         filtered_df = filtered_df[filtered_df[col].isin(selected_values)]
#
#     # Display filtered DataFrame and total quantity in separate columns
#     col1, col2 = st.columns(2)
#     with col1:
#         st.write(filtered_df)
#
#     with col2:
#         total_of_material = filtered_df['وزن البوليصة '].sum()
#         st.dataframe({'اجمالى الكمية الموردة': [total_of_material]}, width=400)


import pandas as pd
import openpyxl
import streamlit as st

st.set_page_config(layout='wide')

# Load Excel sheet
FILEPATH = "اجمالى الحديد والأسمنت.xlsx"
work_book = openpyxl.load_workbook(FILEPATH)
sheet_names = work_book.sheetnames

st.sidebar.header('واردات الحديد والأسمنت')

type_of_imports = st.sidebar.selectbox('اختر نوع التوريد', options=sheet_names)

if type_of_imports in sheet_names:
    df = pd.read_excel(FILEPATH, sheet_name=type_of_imports, skiprows=4)
    column_names = df.columns.tolist()
    filter_columns = st.multiselect('اختار العمود الذى تريد عمل فلتر عليه', options=column_names)

    filtered_df = df.copy()
    for col in filter_columns:
        selected_values = st.multiselect(f'اختر {col}', options=filtered_df[col].unique())
        filtered_df = filtered_df[filtered_df[col].isin(selected_values)]

    # Display filtered DataFrame and total quantity in separate columns
    col1, col2 = st.columns(2)
    with col1:
        st.dataframe(filtered_df.style.set_properties(**{'text-align': 'center'}), height=600)

    with col2:
        total_of_material = round(filtered_df['وزن البوليصة '].sum(), 2)
        st.dataframe(pd.DataFrame({'اجمالى الكمية الموردة': [total_of_material]}).style.set_properties
                     (**{'text-align': 'center'}), width=400)


